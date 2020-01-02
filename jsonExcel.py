import json
from openpyxl import load_workbook

wb = load_workbook(filename="CombinedReport-Draft.xlsx")

sheet = wb.get_sheet_by_name("Term Report")
sheet2 = wb.get_sheet_by_name("LastLogon")

users = {}

# Using the values_only because you want to return the cells' values
for row in sheet.iter_rows(min_row=2,
                           values_only=True):
    user_id = row[0]
    product = {
        "UserID": row[1],
        "title": row[8],
    }
    users[user_id] = product

# Using json here to be able to format the output for displaying later
print(json.dumps(users))