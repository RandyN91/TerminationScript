import openpyxl
import datetime
wb = openpyxl.load_workbook(filename = 'CombinedReport-Draft.xlsx')

sheet = wb.get_sheet_by_name("Term Report")
sheet2 = wb.get_sheet_by_name("LastLogon")

userlist = []

for col in sheet['B']:

     userlist.append(col.value)

print("\nUser List Loaded\n")



for row in sheet.rows :

    for cell in row:

        if cell.value in userlist:

          #delta =  datetime.datetime.strptime(sheet.cell(row=cell.row, column=9).value, '%Y %M %D %H:%M:%S') - datetime.datetime.strptime(sheet2.cell(row=cell.row, column=2).value,'%Y %M %D %H:%M:%S')

          print(sheet.cell(row=cell.row, column=3).value, sheet.cell(row=cell.row, column=4).value, sheet.cell(row=cell.row, column=9).value , sheet2.cell(row=cell.row, column=2).value)